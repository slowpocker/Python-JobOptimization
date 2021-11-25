import glob
import os
import openpyxl
import pathlib


excel_path = '/content/drive/MyDrive/テスト/Book1.xlsx'

dir_path = '/content/drive/MyDrive/SQL'
if not os.path.exists(dir_path):
    # ディレクトリが存在しない場合、ディレクトリを作成する
    os.makedirs(dir_path)

# 列座標を（数値）をアルファベットに変換する関数
# 数値→アルファベット
def num2alpha(num):
    if num <= 26:
        return chr(64+num)
    elif num % 26 == 0:
        return num2alpha(num//26-1)+chr(90)
    else:
        return num2alpha(num//26)+chr(64+num % 26)


# ブックの取得
actBook = openpyxl.load_workbook(excel_path)

# lists = []

i = 1
# シート数分ループ
for actSheetName in actBook.sheetnames:
    
    # アクティブシートを取得
    actSheet = actBook[actSheetName]
    # 該当シートの最大行と最大列を取得
    maxRow = actSheet.max_row
    maxCol = actSheet.max_column

    # excelシートの１行を1行のテキストとして保存する配列
    contents = []

    # for 行変数 in シート変数.iter_rows(開始行,終了行,開始列,終了列)
    for row in actSheet.iter_rows(min_row=2, max_row=maxRow, min_col=1, max_col=maxCol):
        # １行分のテキストの宣言と初期化
        text = ''
        # for セル変数 in 行変数
        for cell in row:
            # セルを取得
            cellData = cell.value
            # Noneだった場合、空文字に変換する
            if cellData is None:
                cellData = ''
            # 列座標と行座標を取得する（数値）
            colChart = cell.column
            rowChart = cell.row

            if colChart == 'A' and rowChart == 2:
                text += f' (\'{cellData}\','
            elif colChart == 'A' and rowChart != 2:
                text += f',(\'{cellData}\','
            elif colChart == num2alpha(maxCol):
                text += f'\'{cellData}\') '
            else:
                text += f'\'{cellData}\' ,'

        contents.append(text)

    # ファイルのパスを設定する
    text_path = pathlib.Path(f'{dir_path}/sheet{i}.txt')

    with open(text_path, mode='w', encoding='cp932', newline='\n', errors='ignore') as f:
        for content in contents:
            f.write('%s\n' % content)
    f.close()

    i += 1
    # lists.append(contents)

actBook.close


