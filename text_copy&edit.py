import os
import shutil
import openpyxl
# import pathlib

ori_path = '/content/drive/MyDrive/original/SQL'
dir_path = '/content/drive/MyDrive/SQL'
excel_path = '/content/drive/MyDrive/excel/Book2.xlsx'
# 置換前の文字が記載された行（数値）と、その行の開始列（アルファベット）を代入する
minRow = 2
minCol = 'A'
# 置換後の文字が記載された終了行（数値）と、その行の終了列（アルファベット）を代入する
maxRow = 4
maxCol = 'D'

###################################################
# 列座標をアルファベットから数値に変換する関数
def conv(alphabet):
	total = 0
	for n, i in enumerate(alphabet[::-1]):
		chr_num = ord(i.lower())
		chr_A_num = ord('a')
		total +=  (1 + chr_num - chr_A_num)* (26**n)
	return total
####################################################

##########################################
# ① excelファイルから置換元の文字、置換後の文字の配列データを取得する
# 対象のexcelファイルから最初のシート[0]を取得する
actBook = openpyxl.load_workbook(excel_path)
actSheet = actBook.worksheets[0]
# 該当シートの最大行と最大列を取得
# maxRow = actSheet.max_row
# maxCol = actSheet.max_column
# 作成するファイル情報
original_file_contents = []
lists = []
# for 行変数 in シート変数.iter_rows(開始行,終了行,開始列,終了列)
for row in actSheet.iter_rows(min_row=minRow, max_row=maxRow, min_col=conv(minCol), max_col=conv(maxCol)):
    # excelシートの１行を1行のテキストとして保存する配列
    contents = []
    # for セル変数 in 行変数
    for cell in row:
        if cell.row == minRow:
            original_file_contents.append(cell.value)
        else:
            contents.append(cell.value)
    if contents:
        lists.append(contents)
actBook.close
print(original_file_contents)
print(lists)
##########################################
for list in lists:
    formtitle = list[1]
    path = f'/content/drive/MyDrive/編集後フォルダ/{formtitle}'
    if not os.path.exists(path):
        # ディレクトリが存在しない場合、ディレクトリを作成する
        os.makedirs(path)

    # コピー元ファイルの情報をリストとして取得する
    files = os.listdir(ori_path)
    ori_fileList = [f for f in files if os.path.isfile(os.path.join(ori_path, f))]

    for ori_file in ori_fileList:
        # コピー元ファイルを別フォルダ内にコピーする
        src = f'{ori_path}/{ori_file}'
        dest = f'{path}/{ori_file}'
        shutil.copyfile(src, dest)

        # # コピーしたファイル名を取得する
        # before_file = os.path.basename(dest)

        # コピーしたファイル名の文字置換
        # for i in range(0, len(list) - 1, 1):
        after_file_name = ori_file.replace(original_file_contents[0], list[0])
        os.rename(dest, f'{path}/{after_file_name}')

        # ファイル名に合わせてパスを更新
        dest = f'{path}/{after_file_name}'

        # 書き込み用リストを作成
        contents = []

        # ファイルがxml形式ならutf-8,sql形式ならcp932で読み取り•書き込みをする
        if after_file_name.endswith('.xml'):
            char_code = 'utf-8'
        elif after_file_name.endswith('.sql'):
            char_code = 'cp932'
        else:
            char_code = 'cp932'

        # ファイル情報を1行ごとに分割したリストとして取得して、作成するファイル情報に合わせて文字置換、contentsに代入していく
        with open(dest, mode='r', encoding=char_code, errors='ignore') as f:
            data_lines = f.readlines()
            for data_line in data_lines:
                for i in range(0, len(list) - 1, 1):
                    data_line = data_line.replace(original_file_contents[i], list[i])
                contents.append(data_line)
        f.close()

        # contentsに代入された1行ごとのテキストをファイルに書き込んでいく
        with open(dest, mode='w', encoding=char_code, errors='ignore') as f:
            for content in contents:
                f.write(content)
        f.close()
