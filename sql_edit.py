import os
import shutil
import openpyxl

# コピー元ファイルの情報
ori_workname = ''
ori_formtitle = ''
ori_workcode = ''
ori_id = ''
ori_path = '/content/drive/MyDrive/***'

##########################################
# 作成するファイル情報
lists = []
# ブックの取得
excel_path = '/content/drive/MyDrive/テスト/Book1.xlsx'
actBook = openpyxl.load_workbook(excel_path)
# 最初のシートを取得
actSheet = actBook.worksheets[0]
# 該当シートの最大行と最大列を取得
maxRow = actSheet.max_row
maxCol = actSheet.max_column
# for 行変数 in シート変数.iter_rows(開始行,終了行,開始列,終了列)
for row in actSheet.iter_rows(min_row=2, max_row=maxRow, min_col=1, max_col=maxCol):
    # excelシートの１行を1行のテキストとして保存する配列
    contents = []
    # for セル変数 in 行変数
    for cell in row:
        # セルを取得して''の中に入れる
        text = f'\'{cell.value}\''
    contents.append(text)

lists.append(contents)
actBook.close
##########################################

for list in lists:
    workname = list[0]
    formtitle = list[1]
    workcode = list[2]
    id = list[3]
    path = f'/content/drive/MyDrive/***/SQL/{formtitle}'

    if not os.path.exists(path):
        # ディレクトリが存在しない場合、ディレクトリを作成する
        os.makedirs(path)

    # コピー元ファイルの情報をリストとして取得する
    files = os.listdir(ori_path)
    ori_fileList = [f for f in files if os.path.isfile(
        os.path.join(ori_path, f))]

    for ori_file in ori_fileList:
        # コピー元ファイルを別フォルダ内にコピーする
        src = f'{ori_path}/{ori_file}'
        dest = f'{path}/{ori_file}'
        shutil.copyfile(src, dest)

        # コピーしたファイル名を取得する
        before_file = os.path.basename(dest)

        # コピーしたファイル名の文字置換
        after_file = before_file.replace(ori_formtitle, formtitle).replace(
            ori_workcode, workcode).replace(ori_id, id)
        os.rename(dest, f'{path}/{after_file}')

        # ファイル名に合わせてパスを更新
        dest = f'{path}/{after_file}'

        # 書き込み用リストを作成
        contents = []

        # ファイルがxml形式ならutf-8,sql形式ならcp932で読み取り、書き込みをする
        if after_file.endswith('.xml'):
            char_code = 'utf-8'
        elif after_file.endswith('.sql'):
            char_code = 'cp932'
        else:
            char_code = 'utf-8'

        # ファイル情報を1行ごとに分割したリストとして取得して、作成するファイル情報に合わせて文字置換、contentsに代入していく
        with open(dest, mode='r', encoding=char_code, newline='\n', errors='ignore') as f:
            data_lines = f.readlines()
            for data_line in data_lines:
                data_line = data_line.replace(ori_formtitle, formtitle).replace(
                    ori_workcode, workcode).replace(ori_id, id).replace(ori_workname, workname)
                contents.append(data_line)
        f.close()

        # contentsに代入された1行ごとのテキストをファイルに書き込んでいく
        with open(dest, mode='w', encoding=char_code, newline='\n', errors='ignore') as f:
            for content in contents:
                f.write(content)
        f.close()
