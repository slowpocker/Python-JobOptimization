# excelの文字置換
import glob
import os
import openpyxl

# ①対象ファイルのパス
path = 'path'
# ②対象ファイル種別
fileType = '*.xlsx'

# ③置換対象としたいシート名（特定のシートのみ置換をしたい場合に適用）
# sheetName = ['table_bbb', 'table_ccc']

# ④置換前の文字データ
before_wordlist = ['test_flg']
# ⑤置換後の文字データ
after_wordlist = ['aabbcc']

##########################################
excel_path = '/content/drive/MyDrive/excel/Book2.xlsx'
# 置換元の文字を入力した行と、その行の開始列を代入する
first_row = 2
first_col = 1
# excelファイルから置換元の文字、置換後の文字の配列データを取得する
# 対象のexcelファイルから最初のシート[0]を取得する
actBook = openpyxl.load_workbook(excel_path)
actSheet = actBook.worksheets[0]
# 作成するファイル情報
original_file_contents = []
lists = []
# for 行変数 in シート変数.iter_rows(開始行,終了行,開始列,終了列)
for row in actSheet.iter_rows(min_row=first_row, max_row=actSheet.max_row, min_col=first_col, max_col=actSheet.max_column):
    # excelシートの１行を1行のテキストとして保存する配列
    contents = []
    # for セル変数 in 行変数
    for cell in row:
        if cell.row == first_row:
            original_file_contents.append(cell.value)
        else:
            contents.append(cell.value)
    lists.append(contents)
actBook.close
print(original_file_contents)
print(lists)
##########################################


# 「①対象ファイルのパス」配下にある全てのExcelファイルのパスを出力
books_path = glob.glob(os.path.join(path, fileType))
print(books_path)

# 「①対象ファイルのパス」配下にある「xlsx」ファイル数分ループ
for book in books_path:
    print(book)
    bookFlg = 0
    # ブックの取得
    actBook = openpyxl.load_workbook(book)
    # すべてのシートの特定の文字列を置換する
    for actSheet in actBook:
        for i in range(0,len(before_wordlist) - 1,1):
            actSheet.title = actSheet.title.replace(before_wordlist[i], after_wordlist[i])

    # シート数分ループ
    for actSheetName in actBook.sheetnames:
        count = 0

        # 特定のシートのみ置換をしたい場合に以下を適用し、インデントを調整する
        # シート名の判定(「③置換対象としたいシート名」との比較)
        # if actSheetName in sheetName:

        # アクティブシートを取得
        actSheet = actBook[actSheetName]
        # 該当シートの最大行を取得
        maxRow = actSheet.max_row
        maxCol = actSheet.max_column

        # 置換対象項目列のループ
        # for 行変数 in シート変数.iter_rows(開始行,終了行,開始列,終了列)
        for row in actSheet.iter_rows(min_row=1, max_row=maxRow, min_col=1, max_col=maxCol):
            # for セル変数 in 列変数
            for cell in row:
                for i in range(0,len(before_wordlist) - 1,1):
                # 対象セルに「⑤置換後データ」を設定
                    if cell.value in before_wordlist[i]:
                        cell.value = cell.value.replace(before_wordlist[i], after_wordlist[i])
                        count += 1
                        bookFlg = 1
        print(str(count) + "件置換しました。")

    # ブックを保存
    if bookFlg == 1:
        # ブック変数.save(Excelファイルのパス)
        actBook.save(book)
    else:
        actBook.close
